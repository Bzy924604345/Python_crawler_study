import requests
import google.protobuf.text_format as text_format
import bili_pb2
import openpyxl
import re
import time

url = 'http://api.bilibili.com/x/v2/dm/web/seg.so'
#在此键入你的视频oid，通过修改segment_index可抓不同时段包
params = {
    'type':1,         #弹幕类型
    'oid':271283166,    #cid
    #'pid':810872,     #avid
    'segment_index':1 #弹幕分段
}
resp = requests.get(url,params)
data = resp.content

my_seg = bili_pb2.DmSegMobileReply()
my_seg.ParseFromString(data)

wb = openpyxl.Workbook()
ws = wb.create_sheet("弹幕数据！",0)
title = ["id","视频内弹幕出现时间","弹幕模式","弹幕字号","弹幕颜色","发送者mid的HASH"
             ,"弹幕内容","弹幕发送时间","weight","idStr"]
for col in range(1,11):
    ws.cell(row=1, column = col, value = title[col-1])


for j in my_seg.elems:
    parse_data = text_format.MessageToString(j, as_utf8=True) 
    list_data = parse_data.split("\n")
    list_data.pop()
    result_data = []
    
    for data in list_data:
        result_data.append(data.split(": ")[1])
        print(data.split(": ")[1])
    #毫秒转时分秒
    seconds = int(result_data[1]) / 1000
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    result_data[1] = ("%02d:%02d:%02d" % (h, m, s))

    #时间戳转日期格式
    timeStamp_checkpoint= int(result_data[7])
    timeArray = time.localtime(timeStamp_checkpoint)
    checkpoint = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
    result_data[7] = ("%s"%checkpoint)

    ws.append(result_data)

wb.save("./%s-%s弹幕数据.xlsx"%(params.get('oid'),params.get('segment_index')))

